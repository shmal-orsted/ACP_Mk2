import xlwt
import pandas as pd
import json
import csv
from openpyxl import Workbook

def export_to_excel(df):
    """
    Exporting and formatting output to excel for use in papers, sending to others, presentations etc
    :param df:
    :return:
    """

    for key, value in df['Arizona']['SolarAnywhere']['Monthly Dataframe'][0].items():
        value

    wb = Workbook()  # create Workbook object

    sheets = []  #list to store created sheets
    month_list = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
          'August', 'September', 'October', 'November', 'December']

    for site in df.keys():
        ws = wb.create_sheet(site)  #create sheet
        ws.title = site  #title sheet name of site
        sheets.append(ws)  #append sheet to list

    for sheet in sheets:  # worksheet object iterable
        # put each value in the sheet, organized by source and type of data
        x = 2
        y = 1
        dataset_y, dataset_x = y, x
        current_dataset_y = dataset_y + 1
        for dataset in df[sheet.title].keys():
            sheet.cell(row=current_dataset_y-1, column=dataset_x, value=dataset)  # dataset
            for m in range(len(month_list)):
                sheet.cell (row=current_dataset_y+m, column=dataset_x-1, value=month_list[m])  # add month row labels
            # wb.save("file.xlsx")
            # set dataset x,y value

            for month in df[sheet.title][dataset]["Monthly Dataframe"]:
                current_dataset_x = dataset_x


                for year in month.keys():
                    val = month[year]
                    sheet.cell(row=current_dataset_y, column=current_dataset_x, value=val)
                    current_dataset_x = current_dataset_x + 1
                    # wb.save("file.xlsx")

                current_dataset_y = current_dataset_y + 1
                # wb.save("file.xlsx")

            # add in the confidence interval after placing the monthly data
            extra = 0
            for month_bs in df[sheet.title][dataset]["Bootstrap Data"]:
                # ci labels
                sheet.cell(row=current_dataset_y - len(df[sheet.title][dataset]['Monthly Dataframe']) - 1, column=current_dataset_x, value="High")
                sheet.cell(row=current_dataset_y - len(df[sheet.title][dataset]['Monthly Dataframe']) - 1, column=current_dataset_x + 1, value="Low")
                sheet.cell(row=current_dataset_y - len(df[sheet.title][dataset]['Monthly Dataframe']) - 1, column=current_dataset_x + 2, value="Standard Error")
                # add confidence interval info
                sheet.cell(row=current_dataset_y-len(df[sheet.title][dataset]['Monthly Dataframe'])+extra, column=current_dataset_x, value=month_bs[0]["high"])
                sheet.cell(row=current_dataset_y-len(df[sheet.title][dataset]['Monthly Dataframe'])+extra, column=current_dataset_x+1, value=month_bs[0]["low"])
                sheet.cell(row=current_dataset_y-len(df[sheet.title][dataset]['Monthly Dataframe'])+extra, column=current_dataset_x+2, value=month_bs[0]["standard error"])
                # wb.save("file.xlsx")
                extra=extra+1

            dataset_y = dataset_y + len(df[site][dataset]["Bootstrap Data"])+1
            # wb.save("file.xlsx")
            current_dataset_y = current_dataset_y + 1

    # save file
    wb.save("file.xlsx")

    return


def export_to_json(df):

    # print values to excel, increment and print for each location and confidence interval
    # build df from the confidence intervals to print out
    for site in df.keys():
        for source in df[site].keys():
            # todo error here replacing all values with the first one

            # replace month with the current iterative month
            for x in range(0, 12):
                d = {
                    "high": df[site][source]["Bootstrap Data"][x].confidence_interval.high,
                    "low": df[site][source]["Bootstrap Data"][x].confidence_interval.low,
                    "standard error": df[site][source]["Bootstrap Data"][x].standard_error,
                }
                ci_df = pd.DataFrame(data=d, index=[x])
                df[site][source]["Bootstrap Data"][x] = ci_df
                df[site][source]["Bootstrap Data"][x] = ci_df.to_dict(orient="records")
            # need to convert all dfs to dicts in the output first
            # df[site][source]["Data"] = df[site][source]["Data"].to_dict(orient='records')
            del df[site][source]["Data"]
            df[site][source]["Monthly Dataframe"] = df[site][source]["Monthly Dataframe"].to_dict(orient='records')
            # drop the 8760s to lower datasize


    # export to json file
    j = json.dumps(df, indent=4)
    with open('sample.json', 'w') as f:
        print(j, file=f)

    # export to excel too
    export_to_excel(df)

    return